import pandas as pd
import os
import glob
from pathlib import Path
import msoffcrypto
import io
from datetime import datetime
import logging
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import re

class ExcelMerger:
    def __init__(self, default_password="8888"):
        """
        Initialize Excel Merger
        
        Args:
            default_password (str): Default password to try for protected files
        """
        self.default_password = default_password
        self.setup_logging()
        
        # Define the location words to remove from address columns
        self.location_words_to_remove = ["江苏省", "南京市", "建邺区", "江心洲街道", "江心洲"]
        
        # Define street patterns to remove (street name + number + 号)
        # Add more street names to this list as needed
        self.street_names_to_remove = [
            "星月街",
            "绿水街",
            "滨江街",
            "科技路",
            "文武街",
            "绿水街",
            #"贤夫路",

            # "梧桐街",      # Example: uncomment and add more street names like this
            # "银河路",      # Example: 银河路123号 would be removed
            # "金桂大道",    # Example: 金桂大道88号 would be removed
        ]
        
        # Define word replacements (old_word -> new_word)
        # Add more replacements as needed
        self.word_replacements = {
            "星岛街与中新大道交汇处": "",
            "生态科技岛葡园路(葡园路与夹江大桥交汇处)": "",
            "中新大道": "",
            "胜科星洲府": "",
            "升龙公园道": "",
            "长岛观澜": "",
            "公园道": ""
            # "旧词": "新词",           # Example: replace 旧词 with 新词
            # "错误地名": "正确地名",    # Example: replace incorrect place names
            # "简称": "全称",          # Example: replace abbreviations with full names
        }
        
    def setup_logging(self):
        """Set up logging configuration"""
        # Create logs directory if it doesn't exist
        if not os.path.exists('logs'):
            os.makedirs('logs')
        
        # Set up logging
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        log_filename = f'logs/excel_merger_{timestamp}.log'
        
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(levelname)s - %(message)s',
            handlers=[
                logging.FileHandler(log_filename, encoding='utf-8'),
                logging.StreamHandler()  # Also print to console
            ]
        )
        self.logger = logging.getLogger(__name__)
        self.logger.info(f"Excel Merger started - Log file: {log_filename}")
    
    def is_incomplete_address_format(self, text):
        """
        Check if the address text only contains incomplete formats like:
        - X幢X室 (building + room)
        - X-X (dash format)  
        - X栋X (building format)
        Where X represents numbers
        
        Args:
            text: The address text to check
            
        Returns:
            bool: True if the text only contains these incomplete formats
        """
        if text is None or str(text).strip() == '':
            return False
            
        text = str(text).strip()
        
        # Define patterns for incomplete address formats
        patterns = [
            r'^\d+幢\d+室$',           # X幢X室 format
            r'^\d+-\d+$',              # X-X format  
            r'^\d+栋\d+$',             # X栋X format
            r'^\d+幢\d+室\s*$',        # X幢X室 with possible trailing spaces
            r'^\d+-\d+\s*$',           # X-X with possible trailing spaces
            r'^\d+栋\d+\s*$',          # X栋X with possible trailing spaces
        ]
        
        # Check if text matches any of the incomplete patterns
        for pattern in patterns:
            if re.match(pattern, text):
                self.logger.debug(f"🏠 Detected incomplete address format: '{text}' matches pattern '{pattern}'")
                return True
        
        return False
    
    def clean_address_text(self, text):
        """
        Clean the address text by:
        1. Replacing specified words with their replacements
        2. Removing specified location words  
        3. Removing street number patterns
        
        Args:
            text: The original address text
            
        Returns:
            str: Cleaned text with replacements, location words and street patterns removed
        """
        if text is None or str(text).strip() == '':
            return text
            
        cleaned_text = str(text)
        original_text = cleaned_text
        removed_patterns = []  # Track what patterns were removed
        replaced_words = []    # Track what words were replaced
        
        # Step 1: Replace specified words
        for old_word, new_word in self.word_replacements.items():
            if old_word in cleaned_text:
                cleaned_text = cleaned_text.replace(old_word, new_word)
                replaced_words.append(f"'{old_word}' → '{new_word}'")
                self.logger.debug(f"📝 Replaced word: '{old_word}' → '{new_word}'")
        
        # Step 2: Remove each specified location word
        for word in self.location_words_to_remove:
            cleaned_text = cleaned_text.replace(word, "")
        
        # Step 3: Remove street number patterns for each configured street name
        for street_name in self.street_names_to_remove:
            # Create pattern: street_name + one or more digits + 号
            street_pattern = rf'{re.escape(street_name)}\d+号'
            street_matches = re.findall(street_pattern, cleaned_text)
            
            if street_matches:
                for match in street_matches:
                    cleaned_text = cleaned_text.replace(match, "")
                    removed_patterns.append(match)
                    self.logger.debug(f"🏠 Removed street pattern: '{match}'")
        
        # Step 4: Clean up extra spaces and normalize
        cleaned_text = re.sub(r'\s+', ' ', cleaned_text).strip()
        
        # Log the cleaning action if text was actually changed
        if cleaned_text != original_text:
            change_details = []
            if replaced_words:
                change_details.append(f"replaced: {', '.join(replaced_words)}")
            if removed_patterns:
                change_details.append(f"removed: {', '.join(removed_patterns)}")
            
            detail_info = f" ({'; '.join(change_details)})" if change_details else ""
            self.logger.debug(f"🧹 Processed address: '{original_text}' → '{cleaned_text}'{detail_info}")
        
        return cleaned_text
    
    def should_replace_address(self, address_text):
        """
        Check if address should be replaced based on missing key location words
        
        Args:
            address_text: The address text to check
            
        Returns:
            bool: True if address should be replaced (missing 江苏省 or 南京市)
        """
        if address_text is None or str(address_text).strip() == '':
            return True
            
        text = str(address_text)
        has_jiangsu = "江苏省" in text
        has_nanjing = "南京市" in text
        has_jianye = "建邺区" in text
        has_jiangxinzhou = "江心洲" in text
        
        # Replace if missing either 江苏省 or 南京市
        should_replace = not (has_jiangsu and has_nanjing and has_jianye and has_jiangxinzhou)
        
        if should_replace:
            self.logger.debug(f"🔍 Address needs replacement (missing location info): '{text}'")
        
        return should_replace
    
    def find_column_indices(self, headers):
        """
        Find the indices of key columns in the headers
        
        Args:
            headers (list): List of header values
            
        Returns:
            dict: Dictionary with column indices
        """
        indices = {
            'pickup_point': -1,  # 自提点
            'detailed_address': -1,  # 详细地址
        }
        
        for idx, header in enumerate(headers):
            if header:
                header_str = str(header).strip()
                if "自提点" in header_str:
                    indices['pickup_point'] = idx
                    self.logger.info(f"📍 Found pickup point column '自提点' at index {idx}")
                elif "详细地址" in header_str:
                    indices['detailed_address'] = idx
                    self.logger.info(f"📍 Found detailed address column '详细地址' at index {idx}")
        
        if indices['pickup_point'] == -1:
            self.logger.warning("⚠️ '自提点' column not found in headers")
        if indices['detailed_address'] == -1:
            self.logger.warning("⚠️ '详细地址' column not found in headers")
            
        return indices
    
    def is_password_protected(self, file_path):
        """
        Check if Excel file is password protected
        
        Args:
            file_path (str): Path to Excel file
            
        Returns:
            bool: True if password protected, False otherwise
        """
        try:
            # Try to read without password first
            pd.read_excel(file_path, nrows=0)  # Just read header
            return False
        except Exception as e:
            error_msg = str(e).lower()
            if 'password' in error_msg or 'encrypted' in error_msg or 'protected' in error_msg:
                return True
            # Try with msoffcrypto to detect encryption
            try:
                with open(file_path, 'rb') as file:
                    office_file = msoffcrypto.OfficeFile(file)
                    if office_file.is_encrypted():
                        return True
            except:
                pass
            return False
    
    def read_excel_with_formatting(self, file_path, password=None):
        """
        Read Excel file with formatting preserved using openpyxl
        
        Args:
            file_path (str): Path to the Excel file
            password (str): Password for protected files (optional)
            
        Returns:
            tuple: (openpyxl.Worksheet or None, success_status, error_message)
        """
        filename = os.path.basename(file_path)
        
        try:
            if password:
                # Handle password-protected files
                with open(file_path, 'rb') as file:
                    office_file = msoffcrypto.OfficeFile(file)
                    office_file.load_key(password=password)
                    
                    decrypted = io.BytesIO()
                    
                    # Try both methods for compatibility
                    try:
                        office_file.save(decrypted)  # Older version
                    except AttributeError:
                        office_file.decrypt(decrypted)  # Newer version
                    
                    decrypted.seek(0)
                    workbook = load_workbook(decrypted)
                    worksheet = workbook.active
                    
                    return worksheet, True, None
            else:
                # Handle regular files
                workbook = load_workbook(file_path)
                worksheet = workbook.active
                return worksheet, True, None
                
        except Exception as e:
            error_msg = str(e)
            self.logger.error(f"❌ {filename} - Failed to read with formatting: {error_msg}")
            return None, False, error_msg

    def copy_cell_formatting(self, source_cell, target_cell):
        """
        Copy formatting from source cell to target cell
        """
        if source_cell.font:
            target_cell.font = Font(
                name=source_cell.font.name,
                size=source_cell.font.size,
                bold=source_cell.font.bold,
                italic=source_cell.font.italic,
                underline=source_cell.font.underline,
                strike=source_cell.font.strike,
                color=source_cell.font.color
            )
        
        if source_cell.fill:
            target_cell.fill = PatternFill(
                fill_type=source_cell.fill.fill_type,
                start_color=source_cell.fill.start_color,
                end_color=source_cell.fill.end_color
            )
        
        if source_cell.border:
            target_cell.border = Border(
                left=source_cell.border.left,
                right=source_cell.border.right,
                top=source_cell.border.top,
                bottom=source_cell.border.bottom
            )
        
        if source_cell.alignment:
            target_cell.alignment = Alignment(
                horizontal=source_cell.alignment.horizontal,
                vertical=source_cell.alignment.vertical,
                wrap_text=source_cell.alignment.wrap_text
            )
    
    def is_header_row(self, row):
        """
        Check if a row looks like a header row
        Header rows typically contain text like: 跟团号, 下单人, 团员备注, etc.
        """
        header_keywords = ['跟团号', '下单人', '团员备注', '支付时间', '团长备注', '商品', 
                          '订单金额', '退款金额', '订单状态', '自提点', '详细地址']
        
        row_values = [str(cell.value).strip() if cell.value is not None else '' for cell in row]
        
        # Check if at least 3 header keywords are found in this row
        matching_keywords = sum(1 for keyword in header_keywords if any(keyword in value for value in row_values))
        
        return matching_keywords >= 3
    
    def extract_header_from_row(self, row):
        """
        Extract header values from a detected header row
        
        Args:
            row: openpyxl row object
            
        Returns:
            list: List of header values
        """
        header_values = []
        for cell in row:
            if cell.value is not None:
                header_values.append(str(cell.value).strip())
            else:
                header_values.append('')
        return header_values

    def merge_excel_files_with_formatting(self, folder_path=".", output_file='1.xlsx'):
        """
        Merge multiple Excel files with formatting preserved and process address columns
        
        Args:
            folder_path (str): Path to folder containing Excel files
            output_file (str): Name of output file
        """
        
        self.logger.info(f"📂 Starting formatted merge process in folder: {os.path.abspath(folder_path)}")
        
        # Find all Excel files in the current directory
        excel_extensions = ['*.xlsx', '*.xls', '*.xlsm']
        excel_files = []
        
        for extension in excel_extensions:
            excel_files.extend(glob.glob(os.path.join(folder_path, extension)))
        
        # Remove the output file from the list if it exists
        excel_files = [f for f in excel_files if not f.endswith(output_file)]
        
        if not excel_files:
            self.logger.warning(f"⚠️ No Excel files found in {folder_path}")
            return
        
        self.logger.info(f"📋 Found {len(excel_files)} Excel files to process")
        
        # Create new workbook for output
        output_workbook = Workbook()
        output_worksheet = output_workbook.active
        output_worksheet.title = "Merged Data"
        
        successful_files = []
        failed_files = []
        current_row = 1
        header_added = False
        detected_headers = None  # Store the actual headers from source files
        column_indices = {}  # Store column indices
        
        # Statistics
        total_addresses_replaced = 0
        total_addresses_cleaned = 0
        total_addresses_prepended = 0  # New counter for prepended addresses
        
        # Process each file
        for i, file_path in enumerate(excel_files):
            filename = os.path.basename(file_path)
            self.logger.info(f"📖 Processing file {i+1}/{len(excel_files)}: {filename}")
            
            # Determine if file is password protected
            is_protected = self.is_password_protected(file_path)
            password = self.default_password if is_protected else None
            
            # Read with formatting preserved
            worksheet, success, error = self.read_excel_with_formatting(file_path, password)
            
            if success and worksheet is not None:
                rows_data = list(worksheet.iter_rows())
                
                # Debug logging
                self.logger.info(f"📄 {filename} - Total rows found: {len(rows_data)}")
                
                # Check if worksheet has any rows
                if len(rows_data) == 0:
                    self.logger.warning(f"⚠️ {filename} - File is completely empty, skipping")
                    failed_files.append((filename, "File is completely empty"))
                    continue
                
                # Find header row and data rows
                header_row_index = -1
                data_rows = []
                
                # Look for header row
                for idx, row in enumerate(rows_data):
                    if self.is_header_row(row):
                        header_row_index = idx
                        self.logger.info(f"📄 {filename} - Header found at row {idx + 1}")
                        break
                
                if header_row_index >= 0:
                    # Found header, get data rows after header
                    header_row = rows_data[header_row_index]
                    
                    # Extract and store header values if not done yet
                    if detected_headers is None:
                        detected_headers = self.extract_header_from_row(header_row)
                        column_indices = self.find_column_indices(detected_headers)
                        self.logger.info(f"📄 {filename} - Detected headers: {detected_headers[:5]}...")  # Show first 5
                    
                    potential_data_rows = rows_data[header_row_index + 1:]
                    
                    # Filter non-empty data rows
                    for row in potential_data_rows:
                        row_values = [cell.value for cell in row if cell.value is not None and str(cell.value).strip() != '']
                        if row_values:  # Row has actual data
                            data_rows.append(row)
                    
                    self.logger.info(f"📄 {filename} - Found {len(data_rows)} data rows after header")
                    
                else:
                    # No header found, treat all non-empty rows as data
                    self.logger.info(f"📄 {filename} - No header found, treating all rows as data")
                    for row in rows_data:
                        row_values = [cell.value for cell in row if cell.value is not None and str(cell.value).strip() != '']
                        if row_values:  # Row has actual data
                            data_rows.append(row)
                
                # Add header if not added yet and we have detected headers
                if not header_added and detected_headers is not None:
                    # Create filtered headers (excluding 自提点 column)
                    filtered_headers = []
                    original_to_output_mapping = {}  # Map original column index to output column index
                    output_col_idx = 0
                    
                    for orig_col_idx, header_value in enumerate(detected_headers):
                        # Skip 自提点 column
                        #if orig_col_idx != column_indices.get('pickup_point', -1):
                        #    filtered_headers.append(header_value)
                        #    original_to_output_mapping[orig_col_idx] = output_col_idx
                        #    output_col_idx += 1
                        filtered_headers.append(header_value)
                        original_to_output_mapping[orig_col_idx] = output_col_idx
                        output_col_idx += 1
                    
                    self.column_mapping = original_to_output_mapping
                    
                    # Write the filtered headers
                    for out_col_idx, header_value in enumerate(filtered_headers, 1):
                        target_cell = output_worksheet.cell(row=current_row, column=out_col_idx)
                        target_cell.value = header_value
                        
                        # Apply header formatting if we have the original header row
                        if header_row_index >= 0:
                            # Find the original column index for this header
                            orig_col_idx = -1
                            for orig_idx, mapped_idx in original_to_output_mapping.items():
                                if mapped_idx == out_col_idx - 1:
                                    orig_col_idx = orig_idx
                                    break
                            
                            if orig_col_idx >= 0 and orig_col_idx < len(rows_data[header_row_index]):
                                source_cell = rows_data[header_row_index][orig_col_idx]
                                self.copy_cell_formatting(source_cell, target_cell)
                            else:
                                # Apply basic header formatting
                                target_cell.font = Font(bold=True)
                        else:
                            # Apply basic header formatting
                            target_cell.font = Font(bold=True)
                    
                    current_row += 1
                    header_added = True
                    self.logger.info(f"📄 {filename} - Added header row (keeping all columns including 自提点)")
                
                # Add data rows with enhanced address processing
                if data_rows:
                    file_replaced_count = 0  # Count replaced addresses in this file
                    file_cleaned_count = 0   # Count cleaned addresses in this file  
                    file_prepended_count = 0 # Count prepended addresses in this file
                    
                    for row in data_rows:
                        # Get pickup point and detailed address values for processing
                        pickup_point_value = None
                        detailed_address_value = None
                        
                        if column_indices.get('pickup_point', -1) >= 0 and column_indices['pickup_point'] < len(row):
                            pickup_point_value = row[column_indices['pickup_point']].value
                        
                        if column_indices.get('detailed_address', -1) >= 0 and column_indices['detailed_address'] < len(row):
                            detailed_address_value = row[column_indices['detailed_address']].value
                        
                        # Process detailed address
                        processed_address = detailed_address_value
                        
                        # Step 1: Check if address needs replacement
                        if (column_indices.get('detailed_address', -1) >= 0 and 
                            self.should_replace_address(detailed_address_value) and 
                            pickup_point_value is not None):
                            
                            processed_address = pickup_point_value
                            file_replaced_count += 1
                            total_addresses_replaced += 1
                            self.logger.debug(f"📝 Replaced address with pickup point: '{detailed_address_value}' -> '{pickup_point_value}'")
                        
                        # Step 2: Clean the address (whether original or replaced)
                        if processed_address is not None:
                            original_processed = str(processed_address)
                            cleaned_address = self.clean_address_text(processed_address)
                            
                            if str(cleaned_address) != original_processed:
                                file_cleaned_count += 1
                                total_addresses_cleaned += 1
                                processed_address = cleaned_address
                        
                        # Step 3: NEW - Check if cleaned address only contains incomplete formats
                        # and prepend pickup point if needed
                        if (processed_address is not None and 
                            pickup_point_value is not None and 
                            column_indices.get('detailed_address', -1) >= 0 and
                            self.is_incomplete_address_format(processed_address)):
                            
                            # Clean the pickup point value first
                            cleaned_pickup_point = self.clean_address_text(pickup_point_value)
                            
                            # Prepend cleaned pickup point to the incomplete address
                            original_incomplete = str(processed_address)
                            processed_address = f"{cleaned_pickup_point}{processed_address}"
                            
                            file_prepended_count += 1
                            total_addresses_prepended += 1
                            self.logger.debug(f"🏠 Prepended pickup point to incomplete address: '{original_incomplete}' -> '{processed_address}'")
                        
                        # Write data row (excluding pickup point column)
                        for orig_col_idx, cell in enumerate(row):
                            # Skip pickup point column
                            #if orig_col_idx == column_indices.get('pickup_point', -1):
                            #    continue
                            
                            # Get output column index
                            if orig_col_idx in self.column_mapping:
                                out_col_idx = self.column_mapping[orig_col_idx] + 1  # +1 for 1-based indexing
                                target_cell = output_worksheet.cell(row=current_row, column=out_col_idx)
                                
                                # Use processed address if this is the detailed address column
                                if orig_col_idx == column_indices.get('detailed_address', -1):
                                    target_cell.value = processed_address
                                else:
                                    target_cell.value = cell.value
                                
                                self.copy_cell_formatting(cell, target_cell)
                        
                        current_row += 1
                    
                    successful_files.append(filename)
                    self.logger.info(f"✅ {filename} - Added {len(data_rows)} data rows")
                    
                    if file_replaced_count > 0:
                        self.logger.info(f"📝 {filename} - Replaced {file_replaced_count} addresses with pickup point data")
                    
                    if file_cleaned_count > 0:
                        self.logger.info(f"🧹 {filename} - Cleaned {file_cleaned_count} address entries")
                    
                    if file_prepended_count > 0:
                        self.logger.info(f"🏠 {filename} - Prepended pickup point to {file_prepended_count} incomplete addresses")
                    
                else:
                    self.logger.warning(f"⚠️ {filename} - No data rows found, skipping")
                    failed_files.append((filename, "No data rows found"))
                    
            else:
                failed_files.append((filename, error))
        
        # Auto-adjust column widths
        self.logger.info("📏 Auto-adjusting column widths...")
        for column in output_worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            output_worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Generate summary report
        self.generate_summary_report(successful_files, failed_files, total_addresses_replaced, total_addresses_cleaned, total_addresses_prepended)
        
        if successful_files:
            # Save the formatted workbook
            output_workbook.save(output_file)
            
            total_data_rows = current_row - 2 if header_added else 0  # Subtract header
            self.logger.info(f"🎉 Successfully merged {len(successful_files)} files with formatting preserved")
            self.logger.info(f"📊 Total data rows in merged file: {total_data_rows}")
            self.logger.info(f"📝 Total addresses replaced: {total_addresses_replaced}")
            self.logger.info(f"🧹 Total addresses cleaned: {total_addresses_cleaned}")
            self.logger.info(f"🏠 Total incomplete addresses prepended with pickup point: {total_addresses_prepended}")
            self.logger.info(f"✅ Kept '自提点' column in output")
            self.logger.info(f"💾 Output saved as: {output_file}")
            
        else:
            self.logger.error("❌ No data was successfully read from any files")
        
        output_workbook.close()
    
    def generate_summary_report(self, successful_files, failed_files, total_addresses_replaced=0, total_addresses_cleaned=0, total_addresses_prepended=0):
        """Generate a summary report of the merge process"""
        
        self.logger.info("=" * 70)
        self.logger.info("📊 MERGE SUMMARY REPORT")
        self.logger.info("=" * 70)
        
        self.logger.info(f"✅ Successfully processed files ({len(successful_files)}):")
        if successful_files:
            for i, filename in enumerate(successful_files, 1):
                self.logger.info(f"   {i}. {filename}")
        else:
            self.logger.info("   None")
        
        self.logger.info(f"\n❌ Failed to process files ({len(failed_files)}):")
        if failed_files:
            for i, (filename, error) in enumerate(failed_files, 1):
                self.logger.info(f"   {i}. {filename} - Reason: {error}")
        else:
            self.logger.info("   None")
        
        success_rate = len(successful_files) / (len(successful_files) + len(failed_files)) * 100 if (successful_files or failed_files) else 0
        self.logger.info(f"\n📈 Success Rate: {success_rate:.1f}%")
        
        self.logger.info(f"\n🏠 Address Processing Statistics:")
        self.logger.info(f"   📝 Addresses replaced with pickup point data: {total_addresses_replaced}")
        self.logger.info(f"   🧹 Addresses processed (cleaned): {total_addresses_cleaned}")
        self.logger.info(f"   🏠 Incomplete addresses prepended with pickup point: {total_addresses_prepended}")
        
        if total_addresses_cleaned > 0 or total_addresses_prepended > 0:
            if self.word_replacements:
                replacements_info = [f"'{old}' → '{new}'" for old, new in self.word_replacements.items()]
                self.logger.info(f"   📝 Word replacements configured: {', '.join(replacements_info)}")
            
            self.logger.info(f"   🏷️ Location words removed: {', '.join(self.location_words_to_remove)}")
            
            if self.street_names_to_remove:
                street_patterns = [f"{name}XX号" for name in self.street_names_to_remove]
                self.logger.info(f"   🏠 Street patterns removed: {', '.join(street_patterns)} (XX = numbers)")
        
        self.logger.info(f"\n🏠 Incomplete Address Format Detection:")
        if total_addresses_prepended > 0:
            self.logger.info(f"   ✅ Detected and processed incomplete formats:")
            self.logger.info(f"       • X幢X室 (building + room format)")
            self.logger.info(f"       • X-X (dash format)")
            self.logger.info(f"       • X栋X (building format)")
            self.logger.info(f"   📍 Prepended pickup point data to these incomplete addresses")
        else:
            self.logger.info(f"   ℹ️ No incomplete address formats detected")
        
        self.logger.info(f"\n📋 Column Processing:")
        self.logger.info(f"   ✅ Kept '自提点' column in output")
        self.logger.info(f"   ✅ Processed '详细地址' column with replacement, cleaning, and prepending")
        self.logger.info("=" * 70)

def main():
    """
    Main function to run the Excel merger automatically
    """
    # Configuration - SET YOUR PASSWORD HERE
    DEFAULT_PASSWORD = "8888"  # Change this to your actual password
    OUTPUT_FILE = "1.xlsx"
    
    print("🚀 Automated Excel Files Merger with Enhanced Address Processing")
    print("=" * 70)
    print("This script will:")
    print("- Auto-detect Excel files in current directory")
    print("- Auto-detect password protection")
    print("- Merge all files (excluding headers)")
    print("- Replace incomplete addresses with pickup point data")
    print("- Replace specific words in addresses with correct versions")
    print("- Clean '详细地址' column by removing location words and street patterns")
    print("- Detect incomplete address formats (X幢X室, X-X, X栋X)")
    print("- Prepend pickup point data to incomplete addresses")
    print("- Keep '自提点' column from final output")
    print("- Generate detailed logs")
    print("=" * 70)
    
    # Create merger instance
    merger = ExcelMerger(default_password=DEFAULT_PASSWORD)
    
    # Run the merger with formatting preserved
    merger.merge_excel_files_with_formatting(folder_path=".", output_file=OUTPUT_FILE)
    
    print("\n✨ Process completed! Check the log file for detailed information.")

if __name__ == "__main__":
    # Check required packages
    try:
        import pandas as pd
        import msoffcrypto
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import Font, PatternFill, Border, Alignment
        import re
    except ImportError as e:
        print("Missing required packages. Please install them using:")
        print("pip install pandas openpyxl msoffcrypto-tool")
        print(f"Error: {e}")
        exit(1)
    
    main()